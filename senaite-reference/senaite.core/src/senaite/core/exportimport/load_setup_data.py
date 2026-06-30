# -*- coding: utf-8 -*-
#
# This file is part of SENAITE.CORE.
#
# SENAITE.CORE is free software: you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation, version 2.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.
#
# You should have received a copy of the GNU General Public License along with
# this program; if not, write to the Free Software Foundation, Inc., 51
# Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
# Copyright 2018-2025 by it's authors.
# Some rights reserved, see README and LICENSE.

import tempfile
import traceback

import transaction
from bika.lims import PMF
from bika.lims import api
from bika.lims import logger
from bika.lims.browser import BrowserView
from bika.lims.interfaces import ISetupDataImporter
from openpyxl import load_workbook
from pkg_resources import resource_filename
from Products.CMFCore.utils import getToolByName
from zope.component import getAdapters
from zope.component.hooks import getSite


class LoadSetupData(BrowserView):

    def __init__(self, context, request):
        BrowserView.__init__(self, context, request)
        self.context = context
        self.request = request
        # dependencies to resolve
        self.deferred = []

    def set_reference(self, obj, fieldname, value):
        """Set a (multi-valued) reference field
        """
        field = self.get_field(obj, fieldname)
        if field is None:
            logger.warning(
                "Field '{}' not found on {}".format(fieldname, repr(obj)))
            return
        if self.is_multi_valued(field):
            current = self.get(obj, field)
            if value not in current:
                current.append(value)
                self.set(obj, field, current)
        else:
            self.set(obj, field, value)

    def get_field(self, obj, field_name):
        if api.is_at_content(obj):
            return obj.getField(field_name)
        fields = api.get_fields(obj)
        return fields.get(field_name)

    def is_multi_valued(self, field):
        multiValued = False
        if hasattr(field, "multiValued"):
            multiValued = field.multiValued
        elif hasattr(field, "multi_valued"):
            multiValued = field.multi_valued
        return multiValued

    def get(self, obj, field):
        if api.is_at_content(obj):
            accessor = field.getAccessor(obj)
            return accessor()
        return field.get(obj)

    def set(self, obj, field, value):
        if api.is_at_content(obj):
            mutator = field.getMutator(obj)
            mutator(value)
        else:
            field.set(obj, value)

    def solve_deferred(self, deferred=None):
        # walk through self.deferred, linking ReferenceFields as we go
        unsolved = []
        deferred = deferred if deferred else self.deferred

        for d in deferred:
            src_obj = d["src_obj"]
            src_fieldname = d["src_field"]
            catalog = api.get_tool(d["dest_catalog"])
            try:
                results = catalog(d["dest_query"])
            except Exception:
                continue
            if len(results) > 0:
                value = results[0].getObject()
                self.set_reference(src_obj, src_fieldname, value)
            else:
                unsolved.append(d)

        self.deferred = unsolved
        return len(unsolved)

    def __call__(self):
        form = self.request.form
        portal = getSite()
        workbook = None

        if "setupexisting" in form and "existing" in form and form["existing"]:
            fn = form["existing"].split(":")
            self.dataset_project = fn[0]
            self.dataset_name = fn[1]
            path = "setupdata/%s/%s.xlsx" % \
                (self.dataset_name, self.dataset_name)
            filename = resource_filename(self.dataset_project, path)
            try:
                workbook = load_workbook(filename=filename)  # , use_iterators=True)
            except AttributeError:
                print("")
                print(traceback.format_exc())
                print("Error while loading ", path)

        elif "setupfile" in form and "file" in form and form["file"] and "projectname" in form and form["projectname"]:
            self.dataset_project = form["projectname"]
            tmp = tempfile.mktemp(suffix=".xlsx")
            file_content = form["file"].read()
            open(tmp, "wb").write(file_content)
            workbook = load_workbook(filename=tmp)  # , use_iterators=True)
            self.dataset_name = "uploaded"

        if not workbook:
            message = PMF("File not found...")
            self.context.plone_utils.addPortalMessage(message)
            self.request.RESPONSE.redirect(portal.absolute_url() + "/import")
            return

        adapters = [[name, adapter]
                    for name, adapter
                    in list(getAdapters((self.context, ), ISetupDataImporter))]
        for sheetname in workbook.sheetnames:
            transaction.savepoint()
            ad_name = sheetname.replace(" ", "_")
            if ad_name in [a[0] for a in adapters]:
                adapter = [a[1] for a in adapters if a[0] == ad_name][0]
                adapter(self, workbook, self.dataset_project, self.dataset_name)
                adapters = [a for a in adapters if a[0] != ad_name]
        for name, adapter in adapters:
            transaction.savepoint()
            adapter(self, workbook, self.dataset_project, self.dataset_name)

        check = len(self.deferred)
        while len(self.deferred) > 0:
            new = self.solve_deferred()
            logger.info("solved %s of %s deferred references" % (
                check - new, check))
            if new == check:
                raise Exception("%s unsolved deferred references: %s" % (
                    len(self.deferred), self.deferred))
            check = new

        logger.info("Rebuilding senaite_catalog_setup")
        bsc = getToolByName(self.context, 'senaite_catalog_setup')
        bsc.clearFindAndRebuild()
        logger.info("Rebuilding senaite_catalog")
        bc = getToolByName(self.context, 'senaite_catalog')
        bc.clearFindAndRebuild()
        logger.info("Rebuilding senaite_catalog_analysis")
        bac = getToolByName(self.context, 'senaite_catalog_analysis')
        bac.clearFindAndRebuild()
